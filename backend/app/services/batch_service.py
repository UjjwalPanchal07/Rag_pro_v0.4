import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from app.services.retrieval_service import search_answer_by_module

# Red fill for unanswered cells
_RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_RED_FONT = Font(color="CC0000", bold=True)

NO_ANSWER_MARKER = "No Answer Found"


def process_batch(input_path: str, run_dir: str):
    """
    Reads input.xlsx (No | Question | RFP Level Tag | Module | Answer),
    fills the Answer column by searching MongoDB,
    marks unanswered cells with a red background in the output Excel.

    Returns (output_path, total_questions).
    """
    df = pd.read_excel(input_path)
    df.columns = [c.strip() for c in df.columns]

    # ── Flexible column detection ──────────────────────────────────────────
    col_map = {}
    for col in df.columns:
        low = col.lower()
        if ("question" in low or "query" in low) and "question" not in col_map:
            col_map["question"] = col
        if ("rfp level" in low or "rfp_level" in low or "tag" in low) and "rfp_level_tag" not in col_map:
            col_map["rfp_level_tag"] = col
        if "module" in low and "module" not in col_map:
            col_map["module"] = col
        if ("answer" in low or "response" in low or "reply" in low) and "answer" not in col_map:
            col_map["answer"] = col

    missing = [k for k in ["question", "rfp_level_tag", "module"] if k not in col_map]
    if missing:
        raise ValueError(
            f"Excel missing required columns: {missing}. "
            f"Expected: No, Question, RFP Level Tag, Module, Answer. "
            f"Found: {list(df.columns)}"
        )

    q_col   = col_map["question"]
    tag_col = col_map["rfp_level_tag"]
    mod_col = col_map["module"]

    ans_col = col_map.get("answer", "Answer")
    if ans_col not in df.columns:
        df[ans_col] = ""

    # ── Fill answers row by row ────────────────────────────────────────────
    answers      = []
    unanswered   = []   # track 0-based row indices with no answer

    for idx, (_, row) in enumerate(df.iterrows()):
        question      = str(row[q_col]).strip()
        rfp_level_tag = str(row[tag_col]).strip()
        module        = str(row[mod_col]).strip()

        if not question or question.lower() == "nan":
            answers.append("")
            continue

        answer = search_answer_by_module(rfp_level_tag, module, question)

        if answer:
            answers.append(answer)
        else:
            answers.append(NO_ANSWER_MARKER)
            unanswered.append(idx)

    df[ans_col] = answers

    # ── Save initial Excel via pandas ──────────────────────────────────────
    output_path = os.path.join(run_dir, "output.xlsx")
    df.to_excel(output_path, index=False)

    # ── Apply red highlighting to unanswered cells via openpyxl ───────────
    if unanswered:
        wb  = load_workbook(output_path)
        ws  = wb.active

        # Find the answer column letter in the worksheet (1-indexed, header = row 1)
        header_row  = [cell.value for cell in ws[1]]
        try:
            ans_col_idx = header_row.index(ans_col) + 1   # openpyxl is 1-indexed
        except ValueError:
            ans_col_idx = None

        if ans_col_idx:
            for row_idx in unanswered:
                # row_idx is 0-based pandas index; +2 because Excel row 1 = header
                excel_row = row_idx + 2
                cell = ws.cell(row=excel_row, column=ans_col_idx)
                cell.fill = _RED_FILL
                cell.font = _RED_FONT

        wb.save(output_path)

    answered   = len(answers) - len(unanswered)
    print(f"[batch_service] Done. {len(df)} rows | {answered} answered | {len(unanswered)} unanswered (highlighted red)")
    return output_path, len(df)
